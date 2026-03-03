"""
SACCO Uptime Report Generator
A professional tool for analyzing SACCO transaction uptime reports
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import warnings
from dotenv import load_dotenv
import os
import mysql.connector
from mysql.connector import Error
import hashlib
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from typing import Optional, Dict, List, Tuple
import time
import threading

warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================

st.set_page_config(
    page_title="SACCO Uptime Report Generator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better UI
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #0B5D3B 0%, #1a8c4e 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
        transition: transform 0.3s ease;
    }
    .metric-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    }
    .badge-success { background: #d4edda; color: #155724; padding: 0.25rem 0.5rem; border-radius: 20px; font-size: 0.8rem; font-weight: 600; }
    .badge-warning { background: #fff3cd; color: #856404; padding: 0.25rem 0.5rem; border-radius: 20px; font-size: 0.8rem; font-weight: 600; }
    .badge-danger  { background: #f8d7da; color: #721c24; padding: 0.25rem 0.5rem; border-radius: 20px; font-size: 0.8rem; font-weight: 600; }
    @keyframes fadeIn { from { opacity: 0; transform: translateY(20px); } to { opacity: 1; transform: translateY(0); } }
    .fade-in { animation: fadeIn 0.5s ease-out; }
    .stProgress > div > div > div > div { background-color: #0B5D3B; }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# DATA MODELS
# ============================================================================

@dataclass
class User:
    id: int
    username: str
    email: str
    full_name: str
    role: str
    created_at: datetime
    last_login: Optional[datetime]

@dataclass
class SaccoMetrics:
    name: str
    total_transactions: int
    total_approved: int
    total_customer_related: int
    total_errors: Dict[str, int]
    approval_rate: float
    avg_sigma: float
    defects_per_million: float

@dataclass
class DocumentInfo:
    name: str
    content: bytes
    upload_date: str
    uploaded_by: str
    file_size: int
    record_count: Optional[int] = None
    document_id: Optional[int] = None

# ============================================================================
# DATABASE MANAGER
# ============================================================================

class DatabaseManager:
    """
    Handles all database operations with connection pooling and async logging
    """

    DB_CONNECT_TIMEOUT = 5   # seconds for TCP handshake
    THREAD_JOIN_TIMEOUT = 6  # seconds to wait for the thread

    def __init__(self):
        load_dotenv()
        self.config = {
            'host':     os.getenv('DB_HOST', 'localhost'),
            'user':     os.getenv('DB_USER', ''),
            'password': os.getenv('DB_PASSWORD', ''),
            'database': os.getenv('DB_NAME', ''),
        }

    # ------------------------------------------------------------------
    # Connection management
    # ------------------------------------------------------------------
    def connect_with_timeout(self):
        """
        Returns a live connection or None.
        Uses a background thread so a stalled TCP connect can never
        block Streamlit's main thread beyond THREAD_JOIN_TIMEOUT seconds.
        """
        conn_holder = [None]
        err_holder  = [None]

        def _connect():
            try:
                conn_holder[0] = mysql.connector.connect(
                    **self.config,
                    connection_timeout=self.DB_CONNECT_TIMEOUT,
                    use_pure=True,
                )
            except Exception as exc:
                err_holder[0] = exc

        t = threading.Thread(target=_connect, daemon=True)
        t.start()
        t.join(timeout=self.THREAD_JOIN_TIMEOUT)

        if t.is_alive():
            st.error("⏱️ Database connection timed out. Check your DB_HOST / network settings.")
            return None

        if err_holder[0]:
            st.error(f"Database connection failed: {err_holder[0]}")
            return None

        return conn_holder[0]

    def get_connection(self):
        return self.connect_with_timeout()

    # ------------------------------------------------------------------
    # Database initialization
    # ------------------------------------------------------------------
    def init_database(self) -> bool:
        """
        Creates tables and default admin user if they don't exist.
        Called once and the result is cached in session_state.
        """
        if st.session_state.get('db_initialized'):
            return True

        conn = self.get_connection()
        if not conn:
            return False

        try:
            cursor = conn.cursor()

            # users table
            cursor.execute("SHOW TABLES LIKE 'users'")
            if not cursor.fetchone():
                cursor.execute("""
                    CREATE TABLE users (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        username VARCHAR(50) UNIQUE NOT NULL,
                        password_hash VARCHAR(255) NOT NULL,
                        email VARCHAR(100),
                        full_name VARCHAR(100),
                        role VARCHAR(20) DEFAULT 'user',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        last_login TIMESTAMP NULL,
                        password_changed_at TIMESTAMP NULL,
                        is_active BOOLEAN DEFAULT TRUE
                    )
                """)
            else:
                cursor.execute("SHOW COLUMNS FROM users LIKE 'is_active'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE users ADD COLUMN is_active BOOLEAN DEFAULT TRUE")

            # audit_log table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS audit_log (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT,
                    action VARCHAR(50),
                    details TEXT,
                    ip_address VARCHAR(45),
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            """)

            # uploaded_documents table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS uploaded_documents (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    filename VARCHAR(255),
                    file_hash VARCHAR(64),
                    uploaded_by INT,
                    upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    record_count INT,
                    FOREIGN KEY (uploaded_by) REFERENCES users(id)
                )
            """)

            # default admin user
            cursor.execute("SELECT id FROM users WHERE username = 'admin'")
            if not cursor.fetchone():
                password_hash = hashlib.sha256('admin123'.encode()).hexdigest()
                cursor.execute("""
                    INSERT INTO users (username, password_hash, role, full_name, email, is_active)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, ('admin', password_hash, 'admin', 'System Administrator',
                      'admin@system.com', True))

            conn.commit()
            st.session_state['db_initialized'] = True
            return True

        except Error as e:
            st.error(f"Database initialization error: {e}")
            return False
        finally:
            cursor.close()
            conn.close()

    # ------------------------------------------------------------------
    # User management
    # ------------------------------------------------------------------
    def verify_user(self, username: str, password: str) -> Optional[User]:
        conn = self.get_connection()
        if not conn:
            return None

        try:
            cursor = conn.cursor(dictionary=True)
            password_hash = hashlib.sha256(password.encode()).hexdigest()

            cursor.execute("""
                UPDATE users
                SET last_login = NOW()
                WHERE username = %s AND password_hash = %s AND is_active = TRUE
            """, (username, password_hash))

            if cursor.rowcount == 0:
                conn.rollback()
                return None

            conn.commit()

            cursor.execute("""
                SELECT id, username, email, full_name, role, created_at, last_login
                FROM users WHERE username = %s
            """, (username,))

            user_data = cursor.fetchone()
            if not user_data:
                return None

            self._log_async(user_data['id'], 'LOGIN', 'User logged in')

            return User(
                id=user_data['id'],
                username=user_data['username'],
                email=user_data['email'],
                full_name=user_data['full_name'],
                role=user_data['role'],
                created_at=user_data['created_at'],
                last_login=user_data['last_login'],
            )

        except Error as e:
            st.error(f"Verification error: {e}")
            return None
        finally:
            cursor.close()
            conn.close()

    def create_user(self, username: str, password: str, email: str,
                    full_name: str, role: str = 'user') -> bool:
        conn = self.get_connection()
        if not conn:
            return False
        try:
            cursor = conn.cursor()
            password_hash = hashlib.sha256(password.encode()).hexdigest()
            cursor.execute("""
                INSERT INTO users (username, password_hash, email, full_name, role, is_active)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (username, password_hash, email, full_name, role, True))
            conn.commit()
            return True
        except Error as e:
            st.error(f"User creation error: {e}")
            return False
        finally:
            cursor.close()
            conn.close()

    def get_all_users(self) -> List[Dict]:
        conn = self.get_connection()
        if not conn:
            return []
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT id, username, email, full_name, role, created_at, last_login, is_active
                FROM users ORDER BY created_at DESC
            """)
            return cursor.fetchall()
        except Error as e:
            st.error(f"Error fetching users: {e}")
            return []
        finally:
            cursor.close()
            conn.close()

    def delete_user(self, user_id: int) -> bool:
        conn = self.get_connection()
        if not conn:
            return False
        try:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE users SET is_active = FALSE
                WHERE id = %s AND username != 'admin'
            """, (user_id,))
            conn.commit()
            return cursor.rowcount > 0
        except Error as e:
            st.error(f"User deletion error: {e}")
            return False
        finally:
            cursor.close()
            conn.close()

    def change_password(self, username: str, old_password: str, new_password: str) -> bool:
        conn = self.get_connection()
        if not conn:
            return False
        try:
            cursor = conn.cursor()
            old_hash = hashlib.sha256(old_password.encode()).hexdigest()
            new_hash = hashlib.sha256(new_password.encode()).hexdigest()
            cursor.execute("""
                UPDATE users
                SET password_hash = %s, password_changed_at = NOW()
                WHERE username = %s AND password_hash = %s
            """, (new_hash, username, old_hash))
            success = cursor.rowcount > 0
            conn.commit()
            if success:
                cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
                row = cursor.fetchone()
                if row:
                    self._log_async(row[0], 'PASSWORD_CHANGE', 'Password changed')
            return success
        except Error as e:
            st.error(f"Password change error: {e}")
            return False
        finally:
            cursor.close()
            conn.close()

    # ------------------------------------------------------------------
    # Document management
    # ------------------------------------------------------------------
    def save_document(self, filename: str, content: bytes, uploaded_by: str, record_count: int = None) -> bool:
        """Save document metadata to database"""
        conn = self.get_connection()
        if not conn:
            return False
        
        try:
            cursor = conn.cursor()
            
            # Get user ID
            cursor.execute("SELECT id FROM users WHERE username = %s", (uploaded_by,))
            user_result = cursor.fetchone()
            if not user_result:
                return False
            user_id = user_result[0]
            
            # Calculate file hash for uniqueness
            file_hash = hashlib.sha256(content).hexdigest()
            
            # Insert document record
            cursor.execute("""
                INSERT INTO uploaded_documents (filename, file_hash, uploaded_by, record_count)
                VALUES (%s, %s, %s, %s)
            """, (filename, file_hash, user_id, record_count))
            
            conn.commit()
            return True
            
        except Error as e:
            st.error(f"Error saving document: {e}")
            return False
        finally:
            cursor.close()
            conn.close()

    def get_all_documents(self) -> List[Dict]:
        """Get all uploaded documents from database"""
        conn = self.get_connection()
        if not conn:
            return []
        
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT d.id, d.filename, d.file_hash, d.upload_date, d.record_count,
                       u.username as uploaded_by
                FROM uploaded_documents d
                JOIN users u ON d.uploaded_by = u.id
                ORDER BY d.upload_date DESC
            """)
            return cursor.fetchall()
            
        except Error as e:
            st.error(f"Error fetching documents: {e}")
            return []
        finally:
            cursor.close()
            conn.close()

    def delete_document(self, doc_id: int) -> bool:
        """Delete a document from database"""
        conn = self.get_connection()
        if not conn:
            return False
        
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM uploaded_documents WHERE id = %s", (doc_id,))
            conn.commit()
            return cursor.rowcount > 0
        except Error as e:
            st.error(f"Error deleting document: {e}")
            return False
        finally:
            cursor.close()
            conn.close()

    # ------------------------------------------------------------------
    # Async logging
    # ------------------------------------------------------------------
    def _log_async(self, user_id: int, action: str, details: str):
        threading.Thread(
            target=self.log_action,
            args=(user_id, action, details),
            daemon=True,
        ).start()

    def log_action(self, user_id: int, action: str, details: str):
        try:
            conn = mysql.connector.connect(
                **self.config,
                connection_timeout=self.DB_CONNECT_TIMEOUT,
                use_pure=True,
            )
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO audit_log (user_id, action, details) VALUES (%s, %s, %s)",
                (user_id, action, details),
            )
            conn.commit()
            cursor.close()
            conn.close()
        except Exception:
            pass

# ============================================================================
# DATA PROCESSOR
# ============================================================================

class DataProcessor:
    REQUIRED_COLUMNS = ['BANK', 'TRX_DATE', 'APPROVED', 'CUSTOMER RELATED',
                        'TIME OUT ERROR 911', 'UNREACHABLE 912', 'BANK SYSTEM_ERROR 909']

    @staticmethod
    def process_uptime_data(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        header_row = DataProcessor._find_header_row(df)
        if header_row is not None:
            df.columns = df.iloc[header_row].astype(str).str.strip()
            df = df.iloc[header_row + 1:].reset_index(drop=True)
        df = DataProcessor._clean_dataframe(df)
        df = DataProcessor._convert_data_types(df)
        df = DataProcessor._filter_valid_rows(df)
        return df

    @staticmethod
    def _find_header_row(df: pd.DataFrame) -> Optional[int]:
        for idx, row in df.iterrows():
            if isinstance(row.iloc[0], str) and 'BANK' in str(row.iloc[0]).upper():
                return idx
        return None

    @staticmethod
    def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        df = df.dropna(how='all').dropna(axis=1, how='all')
        df.columns = df.columns.str.strip()
        return df

    @staticmethod
    def _convert_data_types(df: pd.DataFrame) -> pd.DataFrame:
        if 'TRX_DATE' in df.columns:
            df['TRX_DATE'] = pd.to_datetime(df['TRX_DATE'], errors='coerce')
        for col in ['APPROVED', 'CUSTOMER RELATED', 'TIME OUT ERROR 911',
                    'UNREACHABLE 912', 'BANK SYSTEM_ERROR 909']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        for col in ['APPROVAL RATE', 'SIGMA', 'DEFECTS PER 1M']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        return df

    @staticmethod
    def _filter_valid_rows(df: pd.DataFrame) -> pd.DataFrame:
        df = df[df['BANK'].notna() & (df['BANK'].astype(str).str.strip() != '')]
        df = df[~df['BANK'].astype(str).str.upper().str.contains('TOTAL|AVERAGE|SUM', na=False)]
        return df

    @staticmethod
    def generate_sacco_report(df: pd.DataFrame, sacco_name: str) -> Optional[Tuple[pd.DataFrame, pd.DataFrame, int, float]]:
        sacco_data = df[df['BANK'].astype(str).str.contains(sacco_name, case=False, na=False)].copy()
        if sacco_data.empty:
            return None
        if 'TRX_DATE' in sacco_data.columns:
            sacco_data = sacco_data.sort_values('TRX_DATE')
        metrics = DataProcessor._calculate_metrics(sacco_data)
        summary = DataProcessor._create_summary(metrics, sacco_data)
        return sacco_data, summary, metrics['total_transactions'], metrics['approval_rate']

    @staticmethod
    def _calculate_metrics(sacco_data: pd.DataFrame) -> Dict:
        totals = {
            'approved':     sacco_data['APPROVED'].sum()              if 'APPROVED'              in sacco_data.columns else 0,
            'customer':     sacco_data['CUSTOMER RELATED'].sum()      if 'CUSTOMER RELATED'      in sacco_data.columns else 0,
            'timeout':      sacco_data['TIME OUT ERROR 911'].sum()    if 'TIME OUT ERROR 911'    in sacco_data.columns else 0,
            'unreachable':  sacco_data['UNREACHABLE 912'].sum()       if 'UNREACHABLE 912'       in sacco_data.columns else 0,
            'system_error': sacco_data['BANK SYSTEM_ERROR 909'].sum() if 'BANK SYSTEM_ERROR 909' in sacco_data.columns else 0,
        }
        total_transactions = sum(totals.values())
        approval_rate = (totals['approved'] / total_transactions * 100) if total_transactions > 0 else 0
        return {**totals, 'total_transactions': total_transactions, 'approval_rate': approval_rate}

    @staticmethod
    def _create_summary(metrics: Dict, sacco_data: pd.DataFrame) -> pd.DataFrame:
        return pd.DataFrame({
            'Metric': [
                'Total Approved', 'Customer Related', 'Timeout Error (911)',
                'Unreachable (912)', 'System Error (909)', 'Total Transactions',
                'Overall Approval Rate (%)', 'Average Sigma', 'Average Defects per 1M'
            ],
            'Value': [
                f"{metrics['approved']:,.0f}",
                f"{metrics['customer']:,.0f}",
                f"{metrics['timeout']:,.0f}",
                f"{metrics['unreachable']:,.0f}",
                f"{metrics['system_error']:,.0f}",
                f"{metrics['total_transactions']:,.0f}",
                f"{metrics['approval_rate']:.2f}",
                f"{sacco_data['SIGMA'].mean():.2f}"         if 'SIGMA'          in sacco_data.columns else "N/A",
                f"{sacco_data['DEFECTS PER 1M'].mean():,.0f}" if 'DEFECTS PER 1M' in sacco_data.columns else "N/A",
            ]
        })

    @staticmethod
    def get_all_saccos_summary(df: pd.DataFrame) -> pd.DataFrame:
        saccos = sorted([str(s).strip() for s in df['BANK'].dropna().unique() if str(s).strip()])
        summary_data = []
        for sacco in saccos:
            sacco_df = df[df['BANK'].astype(str).str.contains(sacco, case=False, na=False)]
            metrics = DataProcessor._calculate_metrics(sacco_df)
            summary_data.append({
                'SACCO':               sacco,
                'Total Transactions':  metrics['total_transactions'],
                'Total Approved':      metrics['approved'],
                'Customer Related':    metrics['customer'],
                'Errors':              metrics['timeout'] + metrics['unreachable'] + metrics['system_error'],
                'Approval Rate (%)':   round(metrics['approval_rate'], 2),
                'Avg Sigma':           round(sacco_df['SIGMA'].mean(), 2) if 'SIGMA' in sacco_df.columns else 0,
            })
        return pd.DataFrame(summary_data).sort_values('Approval Rate (%)', ascending=False)


# ============================================================================
# VISUALIZATION ENGINE
# ============================================================================

class VisualizationEngine:
    COLOR_PALETTE = {
        'primary':   '#0B5D3B',
        'secondary': '#1a8c4e',
        'success':   '#28a745',
        'warning':   '#ffc107',
        'danger':    '#dc3545',
        'info':      '#17a2b8',
        'error_1':   '#EB0000',
        'error_2':   '#F8A500',
        'error_3':   '#D000C6',
    }

    @staticmethod
    def create_daily_trends(sacco_data: pd.DataFrame, sacco_name: str) -> go.Figure:
        if 'TRX_DATE' in sacco_data.columns:
            plot_data = sacco_data.sort_values('TRX_DATE')
            x_axis = plot_data['TRX_DATE'].dt.strftime('%d-%b')
        else:
            plot_data = sacco_data.reset_index()
            x_axis = plot_data.index + 1

        fig = go.Figure()
        if 'APPROVAL RATE' in plot_data.columns:
            fig.add_trace(go.Scatter(
                x=x_axis, y=plot_data['APPROVAL RATE'],
                mode='lines+markers', name='Approval Rate',
                line=dict(color=VisualizationEngine.COLOR_PALETTE['primary'], width=3),
                marker=dict(size=8, color=VisualizationEngine.COLOR_PALETTE['primary']),
                hovertemplate='Date: %{x}<br>Approval Rate: %{y:.2f}%<extra></extra>'
            ))
        if 'APPROVED' in plot_data.columns:
            fig.add_trace(go.Bar(
                x=x_axis, y=plot_data['APPROVED'],
                name='Approved Transactions',
                marker_color=VisualizationEngine.COLOR_PALETTE['secondary'],
                opacity=0.6, yaxis='y2',
                hovertemplate='Date: %{x}<br>Approved: %{y:,.0f}<extra></extra>'
            ))
        fig.update_layout(
            title=dict(text=f'{sacco_name} - Daily Performance ({sacco_data["APPROVAL RATE"].mean():.2f}%)',
                       font=dict(size=20, color=VisualizationEngine.COLOR_PALETTE['primary'])),
            xaxis=dict(title='Date', tickangle=-45),
            yaxis=dict(title='Approval Rate (%)', range=[0, 105], gridcolor='lightgray', griddash='dot'),
            yaxis2=dict(title='Transaction Count', overlaying='y', side='right', showgrid=False),
            hovermode='x unified', height=400, template='plotly_white',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
            margin=dict(l=60, r=60, t=80, b=60),
        )
        return fig

    @staticmethod
    def create_error_distribution(sacco_data: pd.DataFrame, sacco_name: str) -> go.Figure:
        error_cols   = ['TIME OUT ERROR 911', 'UNREACHABLE 912', 'BANK SYSTEM_ERROR 909']
        error_labels = ['Timeout Error (911)', 'Unreachable (912)', 'System Error (909)']
        error_colors = [VisualizationEngine.COLOR_PALETTE['error_1'],
                        VisualizationEngine.COLOR_PALETTE['error_2'],
                        VisualizationEngine.COLOR_PALETTE['error_3']]

        error_totals = []
        active_labels = []
        active_colors = []
        
        for col, label, color in zip(error_cols, error_labels, error_colors):
            if col in sacco_data.columns:
                total = sacco_data[col].sum()
                if total > 0:
                    error_totals.append(total)
                    active_labels.append(label)
                    active_colors.append(color)

        fig = go.Figure()
        if error_totals:
            fig.add_trace(go.Pie(
                labels=active_labels, values=error_totals, hole=0.4,
                marker=dict(colors=active_colors, line=dict(color='white', width=2)),
                textinfo='percent',
                textposition='inside',
                textfont=dict(size=13),
                domain=dict(x=[0, 0.55]),
                automargin=True,
                hovertemplate='%{label}<br>Count: %{value:,.0f}<br>Percentage: %{percent}<extra></extra>'
            ))
            fig.update_layout(
                title=dict(text=f'{sacco_name} - Error Distribution', x=0.1,
                          font=dict(size=20, color=VisualizationEngine.COLOR_PALETTE['primary'])),
                height=400, template='plotly_white', showlegend=True,
                legend=dict(
                    orientation='v',
                    x=0.60, y=0.5,
                    xanchor='left', yanchor='middle',
                    font=dict(size=13),
                ),
                margin=dict(l=60, r=120, t=80, b=60),
            )
        else:
            fig.add_annotation(
                text="No errors reported",
                xref="paper", yref="paper",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=24, color='gray')
            )
            fig.update_layout(
                title=dict(text=f'{sacco_name} - Error Distribution', x=0.5,
                          font=dict(size=20, color=VisualizationEngine.COLOR_PALETTE['primary'])),
                height=400, template='plotly_white'
            )
        return fig

    @staticmethod
    def create_performance_comparison(summary_df: pd.DataFrame, metric: str = 'Approval Rate (%)', top_n: int = 10):
        top_df    = summary_df.nlargest(top_n, metric)
        bottom_df = summary_df.nsmallest(top_n, metric)
        fig = go.Figure()
        fig.add_trace(go.Bar(
            name='Top Performers', x=top_df['SACCO'], y=top_df[metric],
            marker_color=VisualizationEngine.COLOR_PALETTE['success'],
            text=top_df[metric].round(1), textposition='outside',
            hovertemplate='SACCO: %{x}<br>Rate: %{y:.2f}%<extra></extra>'
        ))
        fig.add_trace(go.Bar(
            name='Bottom Performers', x=bottom_df['SACCO'], y=bottom_df[metric],
            marker_color=VisualizationEngine.COLOR_PALETTE['danger'],
            text=bottom_df[metric].round(1), textposition='outside',
            hovertemplate='SACCO: %{x}<br>Rate: %{y:.2f}%<extra></extra>'
        ))
        fig.update_layout(
            title=dict(text=f'Top and Bottom {top_n} SACCOs by {metric}', x=0.5, font=dict(size=16)),
            xaxis=dict(title='SACCO', tickangle=-45),
            yaxis=dict(title=metric, range=[0, 105]),
            barmode='group', height=500, template='plotly_white', showlegend=True,
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
            margin=dict(l=60, r=60, t=80, b=100),
        )
        return fig

    @staticmethod
    def create_distribution_chart(summary_df: pd.DataFrame, metric: str = 'Approval Rate (%)'):
        fig = go.Figure()
        fig.add_trace(go.Histogram(
            x=summary_df[metric], nbinsx=20,
            marker_color=VisualizationEngine.COLOR_PALETTE['info'],
            opacity=0.7,
            hovertemplate='Range: %{x}<br>Count: %{y}<extra></extra>'
        ))
        mean_value = summary_df[metric].mean()
        fig.add_vline(x=mean_value, line_dash="dash",
                      line_color=VisualizationEngine.COLOR_PALETTE['primary'],
                      annotation_text=f'Mean: {mean_value:.1f}%', annotation_position="top")
        fig.update_layout(
            title=dict(text=f'Distribution of {metric} Across SACCOs', x=0.5, font=dict(size=16)),
            xaxis=dict(title=metric),
            yaxis=dict(title='Number of SACCOs'),
            height=400, template='plotly_white', bargap=0.1,
            margin=dict(l=60, r=60, t=80, b=60),
        )
        return fig


# ============================================================================
# EXCEL REPORT GENERATOR
# ============================================================================

class ExcelReportGenerator:
    COLUMN_RENAME_MAP = {
        'BANK SYSTEM_ERROR 909': 'DECLINE 909',
        'UNREACHABLE 912':       'DECLINE 912',
        'TIME OUT ERROR 911':    'DECLINE 911',
        'CUSTOMER RELATED':      'CUSTOMER',
        'APPROVED':              'Approved',
        'APPROVAL RATE':         'Approval Rate (%)',
        'SIGMA':                 'Sigma',
        'DEFECTS PER 1M':        'Defects/1M',
    }

    @staticmethod
    def create_report(sacco_data: pd.DataFrame, summary: pd.DataFrame, sacco_name: str) -> BytesIO:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            ExcelReportGenerator._write_daily_data(writer, sacco_data, sacco_name)
            summary.to_excel(writer, sheet_name="Summary", index=False)
            ExcelReportGenerator._format_summary_sheet(writer, summary)
        output.seek(0)
        return output

    @staticmethod
    def _write_daily_data(writer, sacco_data, sacco_name):
        daily_data = sacco_data.copy()
        if 'TRX_DATE' in daily_data.columns:
            daily_data['TRX_DATE'] = daily_data['TRX_DATE'].dt.strftime('%d-%b-%Y')
        daily_data.rename(columns=ExcelReportGenerator.COLUMN_RENAME_MAP, inplace=True)
        total_row = ExcelReportGenerator._create_total_row(daily_data)
        daily_data = pd.concat([daily_data, pd.DataFrame([total_row])], ignore_index=True)
        daily_data.to_excel(writer, sheet_name="Daily Data", index=False)
        ExcelReportGenerator._format_daily_sheet(writer, daily_data)

    @staticmethod
    def _create_total_row(daily_data):
        numeric_cols = ['Approved', 'CUSTOMER', 'DECLINE 911', 'DECLINE 912', 'DECLINE 909']
        total_row = {}
        for col in daily_data.columns:
            if col in numeric_cols:
                total_row[col] = daily_data[col].sum()
            elif col in ('Approval Rate (%)', 'Sigma', 'Defects/1M'):
                total_row[col] = daily_data[col].mean()
            else:
                total_row[col] = "TOTAL"
        return total_row

    @staticmethod
    def _format_daily_sheet(writer, daily_data):
        worksheet = writer.sheets["Daily Data"]
        header_fill = PatternFill(start_color="0B5D3B", end_color="0B5D3B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center      = Alignment(horizontal="center", vertical="center")
        total_fill  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_font  = Font(bold=True)

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill, cell.font, cell.alignment = header_fill, header_font, center

        last_row = worksheet.max_row
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=last_row, column=col)
            cell.fill, cell.font, cell.alignment = total_fill, total_font, center

        numeric_formats = {
            'Approved': '#,##0', 'CUSTOMER': '#,##0',
            'DECLINE 911': '#,##0', 'DECLINE 912': '#,##0', 'DECLINE 909': '#,##0',
            'Approval Rate (%)': '0.00', 'Sigma': '0.00', 'Defects/1M': '#,##0',
        }
        for col_idx, column in enumerate(daily_data.columns, 1):
            if column in numeric_formats:
                for row in range(2, worksheet.max_row):
                    worksheet.cell(row=row, column=col_idx).number_format = numeric_formats[column]

        for col_idx, column in enumerate(daily_data.columns, 1):
            max_length = max(
                len(str(column)),
                max(len(str(worksheet.cell(row=r, column=col_idx).value or ""))
                    for r in range(2, worksheet.max_row + 1))
            )
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        worksheet.freeze_panes = "A2"

    @staticmethod
    def _format_summary_sheet(writer, summary):
        worksheet = writer.sheets["Summary"]
        header_fill = PatternFill(start_color="0B5D3B", end_color="0B5D3B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center      = Alignment(horizontal="center", vertical="center")

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill, cell.font, cell.alignment = header_fill, header_font, center

        for col_idx, column in enumerate(summary.columns, 1):
            max_length = max(
                len(str(column)),
                max(len(str(worksheet.cell(row=r, column=col_idx).value or ""))
                    for r in range(2, worksheet.max_row + 1))
            )
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)


# ============================================================================
# UI COMPONENTS
# ============================================================================

class UIComponents:
    @staticmethod
    def show_header():
        st.markdown("""
        <div class="main-header fade-in">
            <h1 style="margin:0">📊 SACCO Uptime Report Generator</h1>
            <p style="margin:0; opacity:0.9">Professional transaction monitoring and analysis tool</p>
        </div>
        """, unsafe_allow_html=True)

    @staticmethod
    def metric_card(title: str, value: str, delta: Optional[str] = None,
                    icon: str = "📈", color: str = "primary"):
        color_map = {'primary': '#0B5D3B', 'success': '#28a745',
                     'warning': '#ffc107', 'danger': '#dc3545'}
        st.markdown(f"""
        <div class="metric-card">
            <div style="font-size:2rem; color:{color_map.get(color, color_map['primary'])}">{icon}</div>
            <div style="font-size:1.5rem; font-weight:bold">{value}</div>
            <div style="color:#666">{title}</div>
            {f'<div style="font-size:0.9rem; color:{color_map.get(color, color_map["primary"])}">{delta}</div>' if delta else ''}
        </div>
        """, unsafe_allow_html=True)

    @staticmethod
    def show_footer():
        st.markdown("---")
        st.markdown("""
        <div style="text-align:center; color:#666; padding:1rem">
            <p>© 2026 SACCO Uptime Report Generator | SACCO DESK</p>
            <p style="font-size:0.8rem">Version 2.0 | All rights reserved</p>
        </div>
        """, unsafe_allow_html=True)


# ============================================================================
# MAIN APPLICATION
# ============================================================================

class StoredFile:
    def __init__(self, content, name):
        self.name = name
        self._content = content
    def getvalue(self): return self._content
    def read(self): return self._content

def main():
    db_manager     = DatabaseManager()
    data_processor = DataProcessor()
    viz_engine     = VisualizationEngine()
    excel_generator = ExcelReportGenerator()
    ui             = UIComponents()

    # ----------------------------------------------------------------
    # Initialize database once per session
    # ----------------------------------------------------------------
    if not st.session_state.get('db_initialized'):
        with st.spinner("Connecting to database…"):
            db_manager.init_database()
        if not st.session_state.get('db_initialized'):
            st.warning(
                "⚠️ Could not reach the database. "
                "Login and user management are unavailable. "
                "Check your `.env` settings and restart."
            )

    # Session-state defaults
    for key, default in [
        ('authenticated', False),
        ('username',      None),
        ('user_role',     None),
        ('documents',     {}),
        ('current_view',  'reports'),
        ('show_settings', False),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ====================================================================
    # LOGIN
    # ====================================================================
    if not st.session_state.authenticated:
        ui.show_header()

        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown("### 🔐 Secure Login")
            login_tab, create_tab = st.tabs(["Login", "Create Account"])

            with login_tab:
                with st.form("login_form"):
                    username = st.text_input("Username", placeholder="Enter your username")
                    password = st.text_input("Password", type="password", placeholder="Enter your password")
                    col_a, col_b = st.columns(2)
                    with col_a:
                        submit = st.form_submit_button("Login", use_container_width=True)
                    with col_b:
                        st.form_submit_button("Reset", type="secondary", use_container_width=True)

                    if submit:
                        if not username or not password:
                            st.warning("⚠️ Please enter both username and password")
                        else:
                            with st.spinner("Verifying credentials…"):
                                user = db_manager.verify_user(username, password)
                            if user:
                                st.session_state.authenticated  = True
                                st.session_state.username       = username
                                st.session_state.user_role      = user.role
                                st.session_state.user_fullname  = user.full_name
                                st.success("✅ Login successful!")
                                st.rerun()
                            else:
                                st.error("❌ Invalid username or password")
            return
            with create_tab:
                with st.form("create_account_form"):
                    new_username  = st.text_input("Username",         placeholder="Choose a username")
                    new_email     = st.text_input("Email",            placeholder="your@email.com")
                    new_fullname  = st.text_input("Full Name",        placeholder="Enter your full name")
                    new_password  = st.text_input("Password",         type="password", placeholder="Choose a password")
                    confirm_password = st.text_input("Confirm Password", type="password")

                    if st.form_submit_button("Create Account", use_container_width=True):
                        if not all([new_username, new_email, new_fullname, new_password, confirm_password]):
                            st.warning("⚠️ Please fill in all fields")
                        elif new_password != confirm_password:
                            st.error("❌ Passwords do not match")
                        elif len(new_password) < 6:
                            st.error("❌ Password must be at least 6 characters")
                        elif not re.match(r"[^@]+@[^@]+\.[^@]+", new_email):
                            st.error("❌ Invalid email format")
                        else:
                            if db_manager.create_user(new_username, new_password, new_email, new_fullname):
                                st.success("✅ Account created! Please login.")
                            else:
                                st.error("❌ Username already exists")
        st.stop()

    # ====================================================================
    # SIDEBAR
    # ====================================================================
    with st.sidebar:
        st.markdown(f"""
        <div style="text-align:center;padding:1rem;background:linear-gradient(135deg,#0B5D3B 0%,#1a8c4e 100%);border-radius:10px;margin-bottom:1rem">
            <h3 style="color:white;margin:0">👋 Welcome</h3>
            <p style="color:white;opacity:.9;margin:0">{st.session_state.get('user_fullname', st.session_state.username)}</p>
            <p style="color:white;opacity:.7;font-size:.8rem;margin:0">{st.session_state.user_role}</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("### 🧭 Navigation")
        nav_options = {"📊 Reports": "reports", "📁 Documents": "documents",
                       "⚙️ Settings": "settings"}
        for label, value in nav_options.items():
            if st.button(label, use_container_width=True,
                         type="primary" if st.session_state.current_view == value else "secondary"):
                st.session_state.current_view = value
                st.session_state.show_settings = (value == "settings")
                st.rerun()

        st.markdown("---")
        if st.button("🔄 Logout", use_container_width=True):
            for key in ['authenticated', 'username', 'user_role', 'user_fullname']:
                st.session_state[key] = None if key != 'authenticated' else False
            st.rerun()

        st.markdown("### 📁 Document Management")
        doc_source = st.radio("Select source:", ["📤 Upload new", "📂 Select from stored"],
                              label_visibility="collapsed")
        uploaded_file = None

        if doc_source == "📤 Upload new":
            uploaded_file = st.file_uploader("Choose Excel file", type=['xlsx', 'xls'],
                                             key="file_uploader",
                                             help="Upload a monthly uptime report")
            if uploaded_file:
                # Process file to get record count
                try:
                    temp_df = pd.read_excel(BytesIO(uploaded_file.getvalue()), header=None)
                    processed_temp = data_processor.process_uptime_data(temp_df)
                    record_count = len(processed_temp) if not processed_temp.empty else 0
                except:
                    record_count = 0
                
                # Save to database
                if db_manager.save_document(
                    filename=uploaded_file.name,
                    content=uploaded_file.getvalue(),
                    uploaded_by=st.session_state.username,
                    record_count=record_count
                ):
                    doc_key = f"{uploaded_file.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    st.session_state.documents[doc_key] = DocumentInfo(
                        name=uploaded_file.name,
                        content=uploaded_file.getvalue(),
                        upload_date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        uploaded_by=st.session_state.username,
                        file_size=len(uploaded_file.getvalue()),
                        record_count=record_count
                    )
                    st.success("✅ Document saved to database")
                else:
                    st.error("❌ Failed to save document to database")

        else:  # "📂 Select from stored"
            # Load documents from database
            db_documents = db_manager.get_all_documents()
            
            if db_documents:
                # Create display options
                doc_options = {}
                for doc in db_documents:
                    upload_date_str = doc['upload_date'].strftime('%Y-%m-%d %H:%M') if doc['upload_date'] else 'Unknown'
                    display_name = f"{doc['filename']} (Uploaded: {upload_date_str})"
                    doc_options[display_name] = doc
                
                selected_doc_name = st.selectbox(
                    "Select document:",
                    options=list(doc_options.keys()),
                    label_visibility="collapsed"
                )
                
                if selected_doc_name:
                    selected_doc = doc_options[selected_doc_name]
                    
                    # Look for the document in session state
                    found = False
                    for doc_key, doc_info in st.session_state.documents.items():
                        if doc_info.name == selected_doc['filename']:
                            uploaded_file = StoredFile(doc_info.content, doc_info.name)
                            st.info(f"📄 Loaded: {doc_info.name}")
                            st.caption(f"Size: {doc_info.file_size/1024:.1f}KB | Uploaded: {doc_info.upload_date}")
                            if doc_info.record_count:
                                st.caption(f"Records: {doc_info.record_count}")
                            found = True
                            break
                    
                    if not found:
                        st.warning("Document content not found in session. Please re-upload.")
            else:
                st.info("No documents in database. Upload a document first.")

        st.markdown("---")
        st.caption("© 2026 SACCO Uptime Analyzer")

    # ====================================================================
    # MAIN CONTENT
    # ====================================================================
    ui.show_header()

    # ---- Settings -------------------------------------------------------
    if st.session_state.show_settings:
        st.markdown("## ⚙️ System Settings")
        settings_tabs = st.tabs(["🔐 Change Password", "👥 User Management", "📊 System Info"])

        with settings_tabs[0]:
            st.markdown("### Change Password")
            with st.form("change_password_form"):
                old_pwd     = st.text_input("Current Password",     type="password")
                new_pwd     = st.text_input("New Password",         type="password")
                confirm_pwd = st.text_input("Confirm New Password", type="password")
                if st.form_submit_button("Update Password", use_container_width=True):
                    if not all([old_pwd, new_pwd, confirm_pwd]):
                        st.warning("Please fill all fields")
                    elif new_pwd != confirm_pwd:
                        st.error("Passwords don't match")
                    elif len(new_pwd) < 6:
                        st.error("Password must be at least 6 characters")
                    else:
                        if db_manager.change_password(st.session_state.username, old_pwd, new_pwd):
                            st.success("Password updated successfully!")
                        else:
                            st.error("Current password is incorrect")

        with settings_tabs[1]:
            if st.session_state.user_role == 'admin':
                st.markdown("### User Management")
                with st.expander("➕ Create New User"):
                    with st.form("create_user_form"):
                        c1, c2 = st.columns(2)
                        with c1:
                            nu = st.text_input("Username")
                            ne = st.text_input("Email")
                        with c2:
                            nf = st.text_input("Full Name")
                            np_ = st.text_input("Password", type="password")
                        role = st.selectbox("Role", ["user", "admin"])
                        if st.form_submit_button("Create User", use_container_width=True):
                            if db_manager.create_user(nu, np_, ne, nf, role):
                                st.success(f"User {nu} created!")
                                st.rerun()

                st.markdown("### Existing Users")
                for user in db_manager.get_all_users():
                    cols = st.columns([3, 2, 2, 1])
                    with cols[0]:
                        st.write(f"**{user['full_name']}**")
                        st.caption(f"@{user['username']}")
                    with cols[1]:
                        st.write(f"Role: {user['role']}")
                        st.caption(f"Joined: {user['created_at'].strftime('%Y-%m-%d') if user['created_at'] else 'N/A'}")
                    with cols[2]:
                        st.write("🟢 Active" if user['is_active'] else "🔴 Inactive")
                    with cols[3]:
                        if user['username'] != 'admin' and user['is_active']:
                            if st.button("🗑️", key=f"del_{user['id']}"):
                                if db_manager.delete_user(user['id']):
                                    st.success(f"User {user['username']} deleted")
                                    st.rerun()
                    st.divider()
            else:
                st.info("User management is only available for administrators")

        with settings_tabs[2]:
            st.markdown("### System Information")
            st.json({
                "Version": "2.0.0",
                "User": st.session_state.username,
                "Role": st.session_state.user_role,
                "Documents in Session": len(st.session_state.documents),
            })

        if st.button("← Back to Reports", use_container_width=True):
            st.session_state.show_settings = False
            st.session_state.current_view  = 'reports'
            st.rerun()
        st.divider()

    # ---- Reports --------------------------------------------------------
    elif st.session_state.current_view == 'reports':
        if uploaded_file is not None:
            try:
                with st.spinner("🔄 Processing file..."):
                    file_content = uploaded_file.getvalue() if hasattr(uploaded_file, 'getvalue') else uploaded_file
                    df = pd.read_excel(
                        BytesIO(file_content) if isinstance(file_content, bytes) else uploaded_file,
                        header=None
                    )
                    processed_df = data_processor.process_uptime_data(df)

                if processed_df.empty:
                    st.error("❌ No valid data found in file")
                    return

                saccos     = sorted(processed_df['BANK'].dropna().unique())
                summary_df = data_processor.get_all_saccos_summary(processed_df)

                st.markdown("## 📊 Dashboard Overview")
                col1, col2, col3, col4 = st.columns(4)
                with col1: ui.metric_card("Total SACCOs",       str(len(saccos)),                                         icon="🏦")
                with col2: ui.metric_card("Total Transactions",  f"{summary_df['Total Transactions'].sum():,.0f}",        icon="💳")
                with col3: ui.metric_card("Avg Approval Rate",   f"{summary_df['Approval Rate (%)'].mean():.1f}%",        icon="📈")
                with col4: ui.metric_card("Avg Sigma",           f"{summary_df['Avg Sigma'].mean():.2f}",                 icon="⚡")

                with st.expander("📈 View All SACCOs Performance", expanded=False):
                    # Create a copy of the dataframe for display
                    display_summary_df = summary_df.copy()
                    
                    # Format Approval Rate to 2 decimal places
                    if 'Approval Rate (%)' in display_summary_df.columns:
                        display_summary_df['Approval Rate (%)'] = display_summary_df['Approval Rate (%)'].apply(
                            lambda x: f"{x:.2f}" if pd.notna(x) else ""
                        )
                    
                    # Format Avg Sigma to 2 decimal places
                    if 'Avg Sigma' in display_summary_df.columns:
                        display_summary_df['Avg Sigma'] = display_summary_df['Avg Sigma'].apply(
                            lambda x: f"{x:.2f}" if pd.notna(x) else ""
                        )
                    cols_to_style = ['Approval Rate (%)', 'Avg Sigma']

                    for col in cols_to_style:
                        if col in display_summary_df.columns:
                            display_summary_df[col] = (
                                pd.to_numeric(display_summary_df[col], errors='coerce')
                            )
                    st.dataframe(
                        display_summary_df.style.background_gradient(
                            subset=['Approval Rate (%)', 'Avg Sigma'], cmap='RdYlGn'),
                        use_container_width=True, hide_index=True
                    )
                    
                    st.download_button(
                        label="📥 Download Summary CSV",
                        data=summary_df.to_csv(index=False),
                        file_name="sacco_summary.csv",
                        mime="text/csv"
                    )

                col1, col2 = st.columns(2)
                with col1: st.plotly_chart(viz_engine.create_performance_comparison(summary_df), use_container_width=True)
                with col2: st.plotly_chart(viz_engine.create_distribution_chart(summary_df),     use_container_width=True)

                st.divider()
                st.markdown("## 🏦 Individual SACCO Analysis")

                col1, col2 = st.columns([3, 1])
                with col1:
                    selected_sacco = st.selectbox("Select SACCO to analyze:", saccos, key="sacco_selector")
                with col2:
                    if st.button("📊 Refresh", use_container_width=True): st.rerun()

                result = data_processor.generate_sacco_report(processed_df, selected_sacco)
                if result:
                    sacco_data, summary, total_txns, approval_rate = result
                    st.markdown(f"### 📈 {selected_sacco} Performance")

                    cols = st.columns(5)
                    metrics_data = [
                        ("Approved",        f"{sacco_data['APPROVED'].sum():,.0f}", "✅"),
                        ("Customer Related", f"{sacco_data['CUSTOMER RELATED'].sum():,.0f}", "👥"),
                        ("Errors",          f"{sacco_data[['TIME OUT ERROR 911','UNREACHABLE 912','BANK SYSTEM_ERROR 909']].sum().sum():,.0f}", "⚠️"),
                        ("Avg Approval",    f"{sacco_data['APPROVAL RATE'].mean():.2f}%", "📊"),
                        ("Avg Sigma",       f"{sacco_data['SIGMA'].mean():.2f}", "⚡"),
                    ]
                    for col, (label, value, icon) in zip(cols, metrics_data):
                        with col: ui.metric_card(label, value, icon=icon)

                    col1, col2 = st.columns(2)
                    with col1: st.plotly_chart(viz_engine.create_daily_trends(sacco_data, selected_sacco),      use_container_width=True)
                    with col2: st.plotly_chart(viz_engine.create_error_distribution(sacco_data, selected_sacco), use_container_width=True)

                    st.markdown("### 📅 Daily Transaction Details")
                    display_cols = ['BANK', 'TRX_DATE', 'APPROVED', 'CUSTOMER RELATED',
                                    'TIME OUT ERROR 911', 'UNREACHABLE 912',
                                    'BANK SYSTEM_ERROR 909', 'APPROVAL RATE', 'SIGMA', 'DEFECTS PER 1M']
                    available_cols = [c for c in display_cols if c in sacco_data.columns]
                    display_df = sacco_data[available_cols].copy()
                    if 'TRX_DATE' in display_df.columns:
                        display_df['TRX_DATE'] = display_df['TRX_DATE'].dt.strftime('%d-%b-%Y')

                    # Calculate total row FIRST (before formatting)
                    total_row = {}
                    for col in display_df.columns:
                        if col in ['APPROVED', 'CUSTOMER RELATED', 'TIME OUT ERROR 911',
                                'UNREACHABLE 912', 'BANK SYSTEM_ERROR 909']:
                            total_row[col] = display_df[col].sum()
                        elif col in ('APPROVAL RATE', 'SIGMA', 'DEFECTS PER 1M'):
                            total_row[col] = display_df[col].mean()
                        else:
                            total_row[col] = "TOTAL"

                    # NOW format the display values to 2 decimal places
                    if 'APPROVAL RATE' in display_df.columns:
                        display_df['APPROVAL RATE'] = display_df['APPROVAL RATE'].apply(
                            lambda x: f"{x:.2f}" if pd.notna(x) else ""
                        )
                        total_row['APPROVAL RATE'] = f"{total_row['APPROVAL RATE']:.2f}" if total_row['APPROVAL RATE'] != "TOTAL" else "TOTAL"

                    if 'SIGMA' in display_df.columns:
                        display_df['SIGMA'] = display_df['SIGMA'].apply(
                            lambda x: f"{x:.2f}" if pd.notna(x) else ""
                        )
                        total_row['SIGMA'] = f"{total_row['SIGMA']:.2f}" if total_row['SIGMA'] != "TOTAL" else "TOTAL"

                    if 'DEFECTS PER 1M' in display_df.columns:
                        display_df['DEFECTS PER 1M'] = display_df['DEFECTS PER 1M'].apply(
                            lambda x: f"{x:.2f}" if pd.notna(x) else ""
                        )
                        total_row['DEFECTS PER 1M'] = f"{total_row['DEFECTS PER 1M']:.2f}" if total_row['DEFECTS PER 1M'] != "TOTAL" else "TOTAL"

                    # Add total row
                    display_df = pd.concat([display_df, pd.DataFrame([total_row])], ignore_index=True)

                    def highlight_total(row):
                        if row.name == len(display_df) - 1:
                            return ['background-color: #f2f2f2; font-weight: bold'] * len(row)
                        return [''] * len(row)

                    st.dataframe(display_df.style.apply(highlight_total, axis=1),
                                 use_container_width=True, hide_index=True)

                    st.markdown("### ⬇️ Export Options")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        csv_data = sacco_data.copy()
                        if 'TRX_DATE' in csv_data.columns:
                            csv_data['TRX_DATE'] = csv_data['TRX_DATE'].dt.strftime('%d-%b-%Y')
                        st.download_button("📥 Download CSV", csv_data.to_csv(index=False),
                                           f"{selected_sacco.replace(' ','_')}_report.csv", "text/csv",
                                           use_container_width=True)
                    with col2:
                        excel_file = excel_generator.create_report(sacco_data, summary, selected_sacco)
                        st.download_button("📥 Download Excel", excel_file,
                                           f"{selected_sacco.replace(' ','_')}_report.xlsx",
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           use_container_width=True)
                    with col3:
                        if st.button("📑 Generate All Report", use_container_width=True):
                            with st.spinner("Generating comprehensive report..."):
                                output = BytesIO()
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    all_summaries = []
                                    for sacco in saccos:
                                        res = data_processor.generate_sacco_report(processed_df, sacco)
                                        if res:
                                            s_data, s_sum, _, _ = res
                                            all_summaries.append({
                                                'SACCO':             sacco,
                                                'Total Approved':    s_sum.iloc[0, 1],
                                                'Total Transactions':s_sum.iloc[5, 1],
                                                'Approval Rate (%)': s_sum.iloc[6, 1],
                                                'Avg Sigma':         s_sum.iloc[7, 1],
                                            })
                                            sheet_name = sacco[:25].replace('/', '_').replace('*', '').replace('?', '')
                                            temp_df = s_data.copy()
                                            if 'TRX_DATE' in temp_df.columns:
                                                temp_df['TRX_DATE'] = temp_df['TRX_DATE'].dt.strftime('%d-%b-%Y')
                                            temp_df.to_excel(writer, sheet_name=sheet_name, index=False)
                                    if all_summaries:
                                        pd.DataFrame(all_summaries).to_excel(
                                            writer, sheet_name='All_SACCOs_Summary', index=False)
                                output.seek(0)
                                st.download_button("📥 Download Complete Report", output,
                                                   "all_saccos_complete_report.xlsx",
                                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                st.exception(e)
        else:
            st.markdown("""
            <div style="text-align:center;padding:3rem;background:#f8f9fa;border-radius:10px">
                <h2 style="color:#0B5D3B">👋 Welcome to SACCO Uptime Report Generator</h2>
                <p style="color:#666;font-size:1.2rem">Please upload a file or select from stored documents to begin</p>
                <p style="color:#999">Supported format: Excel files (.xlsx, .xls)</p>
            </div>
            """, unsafe_allow_html=True)

    # ---- Documents ------------------------------------------------------
    elif st.session_state.current_view == 'documents':
        st.markdown("## 📁 Document Library")
        
        # Load documents from database
        db_documents = db_manager.get_all_documents()
        
        if db_documents:
            # Create a DataFrame for display
            docs_list = []
            for doc in db_documents:
                docs_list.append({
                    'Filename': doc['filename'],
                    'Uploaded By': doc['uploaded_by'],
                    'Upload Date': doc['upload_date'].strftime('%Y-%m-%d %H:%M') if doc['upload_date'] else 'Unknown',
                    'Records': doc['record_count'] if doc['record_count'] else 'N/A',
                    'ID': doc['id']
                })
            
            docs_df = pd.DataFrame(docs_list)
            display_cols = ['Filename', 'Uploaded By', 'Upload Date', 'Records']
            st.dataframe(docs_df[display_cols], use_container_width=True, hide_index=True)
            
            # Option to delete documents (admin only)
            if st.session_state.user_role == 'admin' and len(docs_df) > 0:
                st.markdown("### 🗑️ Delete Documents")
                doc_to_delete = st.selectbox(
                    "Select document to delete:",
                    options=docs_df['Filename'].tolist()
                )
                
                if st.button("Delete Selected Document", type="primary", use_container_width=True):
                    # Find document ID
                    doc_id = None
                    for doc in db_documents:
                        if doc['filename'] == doc_to_delete:
                            doc_id = doc['id']
                            break
                    
                    if doc_id and db_manager.delete_document(doc_id):
                        # Also remove from session state
                        keys_to_remove = []
                        for key, doc_info in st.session_state.documents.items():
                            if doc_info.name == doc_to_delete:
                                keys_to_remove.append(key)
                        for key in keys_to_remove:
                            del st.session_state.documents[key]
                        
                        st.success(f"Document '{doc_to_delete}' deleted!")
                        st.rerun()
                    else:
                        st.error("Failed to delete document")
        else:
            st.info("No documents uploaded yet. Go to the Reports tab to upload documents.")

    # ---- About ----------------------------------------------------------
    elif st.session_state.current_view == 'about':
        st.markdown("## ℹ️ About This System")
        col1, col2 = st.columns([2, 1])
        with col1:
            st.markdown("""
            ### SACCO Uptime Report Generator — Professional Edition

            A comprehensive tool for analyzing and monitoring SACCO transaction uptime reports.

            **Key Features:**
            - 🔐 Secure authentication system
            - 📊 Real-time data processing and analysis
            - 📈 Interactive visualizations
            - 📑 Professional Excel report generation
            - 💾 Document management system with database persistence
            - 👥 Multi-user support with role-based access

            **Version:** 2.0.0

            **Technologies:** Streamlit · MySQL · Plotly · Pandas · OpenPyXL
            """)
        with col2:
            st.image("https://via.placeholder.com/300x200/0B5D3B/ffffff?text=SACCO+Analytics",
                     use_container_width=True)
        st.divider()
        st.markdown("""
        ### 📋 System Requirements
        - Python 3.8+  ·  MySQL Server 5.7+  ·  Modern web browser

        ### 🔧 Configuration — `.env`
        ```env
        DB_HOST=localhost
        DB_USER=your_user
        DB_PASSWORD=your_password
        DB_NAME=sacco_uptime
        ```

        ### 📞 Support
        Contact the system administrator for technical support or feature requests.
        """)

    ui.show_footer()


if __name__ == "__main__":
    main()